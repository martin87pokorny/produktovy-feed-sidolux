"""
Import vyplněného ceníku od výrobce z Excelu do Airtable Produkty_v2.

Vstup:  data/exchange/cenik_pro_vyrobce_VYPLNENY_*.xlsx (od Lakma PL)
Cíl:    pole 'Cena CZK doporučená', 'Cena EUR doporučená'

Cena s DPH se spočítá z 'Cena bez DPH × (1 + sazba/100)' — Excel formuli
neimportujeme (cache nemusí být zapsaná, pokud výrobce soubor neotevřel
v plnotučném Excelu).

Default --dry-run = OFF (live PATCH). Pro preview spusť --dry-run.

Použití:
    python scripts/import_pricing.py path/to/cenik_VYPLNENY.xlsx
    python scripts/import_pricing.py path/to/cenik.xlsx --dry-run
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
ENV_PATH = ROOT / '.env'

# Pozice sloupců v Excelu (1-based, podle scripts/generate_pricing_excel.py)
COL_KOD = 1            # A
COL_CENA_BEZ_CZK = 6   # F
COL_DPH_CZ = 7         # G  (default 21)
COL_CENA_BEZ_EUR = 9   # I
COL_DPH_SK = 10        # J  (default 23)

HDR_ROW = 4            # hlavička v 4. řádku
DATA_START_ROW = 5     # data od 5. řádku

# Validace headeru — substring matching, ať toleruje drobné úpravy textu
EXPECTED_HEADER_SUBSTRINGS: dict[int, str] = {
    COL_KOD: 'Kód Lakma',
    COL_CENA_BEZ_CZK: 'CZK',
    COL_DPH_CZ: 'DPH CZ',
    COL_CENA_BEZ_EUR: 'EUR',
    COL_DPH_SK: 'DPH SK',
}

DEFAULT_DPH_CZ = 21
DEFAULT_DPH_SK = 23


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    with ENV_PATH.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            k, v = line.split('=', 1)
            env[k.strip()] = v.split('  ')[0].strip()
    return env


def fetch_at_records(token: str, base: str, table: str) -> list[dict]:
    fields = ['Kód Lakma', 'Cena CZK doporučená', 'Cena EUR doporučená',
              'Přidat do feedu']
    records: list[dict] = []
    offset: str | None = None
    while True:
        params: list[tuple[str, str]] = [('pageSize', '100')]
        for f in fields:
            params.append(('fields[]', f))
        if offset:
            params.append(('offset', offset))
        url = (
            f'https://api.airtable.com/v0/{base}/{urllib.parse.quote(table)}'
            f'?{urllib.parse.urlencode(params)}'
        )
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        records.extend(data.get('records', []))
        offset = data.get('offset')
        if not offset:
            break
    return records


def patch_batch(url: str, token: str, batch: list[dict]) -> dict:
    body = json.dumps({'records': batch}).encode()
    req = urllib.request.Request(
        url, data=body, method='PATCH',
        headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
    )
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read())
        except urllib.error.HTTPError as e:
            body_text = e.read().decode('utf-8', errors='replace')
            if e.code == 429 and attempt < 2:
                time.sleep(2 * (attempt + 1))
                continue
            raise RuntimeError(f'HTTP {e.code}: {body_text[:400]}')
    raise RuntimeError('unreachable')


def _to_float(value, default: float | None = None) -> float | None:
    if value is None or value == '':
        return default
    if isinstance(value, str):
        s = value.replace('\xa0', '').replace(' ', '').replace(',', '.').strip()
        if not s:
            return default
        try:
            return float(s)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_xlsx(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    """Vrátí (rows, warnings). rows = [{kod, cena_czk, cena_eur, _row_no}, ...]."""
    wb = load_workbook(xlsx_path, data_only=False)
    if 'Ceník' in wb.sheetnames:
        ws = wb['Ceník']
    else:
        ws = wb.active

    # Header validation
    for col, expected in EXPECTED_HEADER_SUBSTRINGS.items():
        cell = ws.cell(row=HDR_ROW, column=col).value
        if not cell or expected.lower() not in str(cell).lower():
            raise SystemExit(
                f"CHYBA: hlavička sloupce {col} ({chr(64+col)}) nesedí. "
                f"Očekávám obsahovat {expected!r}, dostal {cell!r}."
            )

    rows: list[dict] = []
    warnings: list[str] = []

    for r in range(DATA_START_ROW, ws.max_row + 1):
        kod_raw = ws.cell(row=r, column=COL_KOD).value
        if kod_raw is None or str(kod_raw).strip() == '':
            continue
        kod = str(kod_raw).strip()
        if kod.endswith('.0'):  # když Excel uloží jako float
            kod = kod[:-2]

        cena_bez_czk = _to_float(ws.cell(row=r, column=COL_CENA_BEZ_CZK).value)
        sazba_cz = _to_float(ws.cell(row=r, column=COL_DPH_CZ).value, DEFAULT_DPH_CZ)
        cena_bez_eur = _to_float(ws.cell(row=r, column=COL_CENA_BEZ_EUR).value)
        sazba_sk = _to_float(ws.cell(row=r, column=COL_DPH_SK).value, DEFAULT_DPH_SK)

        if cena_bez_czk is None and cena_bez_eur is None:
            warnings.append(f'řádek {r} ({kod}): chybí cena CZK i EUR — vynecháno')
            continue
        if cena_bez_czk is not None and cena_bez_czk <= 0:
            warnings.append(f'řádek {r} ({kod}): cena CZK <= 0 ({cena_bez_czk}) — vynecháno')
            cena_bez_czk = None
        if cena_bez_eur is not None and cena_bez_eur <= 0:
            warnings.append(f'řádek {r} ({kod}): cena EUR <= 0 ({cena_bez_eur}) — vynecháno')
            cena_bez_eur = None

        cena_s_czk = round(cena_bez_czk * (1 + sazba_cz / 100), 2) if cena_bez_czk is not None else None
        cena_s_eur = round(cena_bez_eur * (1 + sazba_sk / 100), 2) if cena_bez_eur is not None else None

        rows.append({
            'kod': kod,
            'cena_czk': cena_s_czk,
            'cena_eur': cena_s_eur,
            'sazba_cz': sazba_cz,
            'sazba_sk': sazba_sk,
            'row_no': r,
        })

    return rows, warnings


def main() -> int:
    sys.stdout.reconfigure(encoding='utf-8')
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', type=Path, help='Cesta k vyplněnému Excel ceníku')
    ap.add_argument('--dry-run', action='store_true', help='Jen preview, nezapisovat')
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f'CHYBA: {args.xlsx} neexistuje.')
        return 2

    print(f'Excel: {args.xlsx}')
    print(f'Mód:   {"DRY-RUN" if args.dry_run else "LIVE PATCH"}')
    print()

    rows, warnings = parse_xlsx(args.xlsx)
    print(f'Excel: {len(rows)} datových řádků s aspoň jednou cenou')
    if warnings:
        print(f'  {len(warnings)} warning(s):')
        for w in warnings[:10]:
            print(f'    - {w}')
        if len(warnings) > 10:
            print(f'    ... a dalších {len(warnings) - 10}')

    # Detekce duplicit Kód Lakma
    seen: dict[str, int] = {}
    dupes: list[str] = []
    for row in rows:
        if row['kod'] in seen:
            dupes.append(f"{row['kod']} (řádek {row['row_no']} a {seen[row['kod']]})")
        else:
            seen[row['kod']] = row['row_no']
    if dupes:
        print(f'CHYBA: duplicitní Kód Lakma v Excelu:')
        for d in dupes[:10]:
            print(f'  {d}')
        return 2

    env = load_env()
    print('\nNačítám AT...')
    at_records = fetch_at_records(env['AIRTABLE_TOKEN'], env['AIRTABLE_BASE_ID'], env['AIRTABLE_TABLE_NAME'])
    print(f'  {len(at_records)} záznamů')

    by_kod = {r['fields'].get('Kód Lakma'): r for r in at_records if r['fields'].get('Kód Lakma')}

    # Match
    not_in_at = [r['kod'] for r in rows if r['kod'] not in by_kod]
    if not_in_at:
        print(f'\n⚠ {len(not_in_at)} Kódů Lakma z Excelu není v AT (vynechány):')
        for kod in not_in_at[:10]:
            print(f'  - {kod}')

    patches: list[dict] = []
    new_czk = overwrite_czk = same_czk = 0
    new_eur = overwrite_eur = same_eur = 0
    samples: list[tuple[str, str, str]] = []

    for row in rows:
        rec = by_kod.get(row['kod'])
        if not rec:
            continue
        cur_czk = rec['fields'].get('Cena CZK doporučená')
        cur_eur = rec['fields'].get('Cena EUR doporučená')

        new_fields: dict = {}

        if row['cena_czk'] is not None:
            if cur_czk is None:
                new_fields['Cena CZK doporučená'] = row['cena_czk']; new_czk += 1
            elif abs(float(cur_czk) - row['cena_czk']) > 0.005:
                new_fields['Cena CZK doporučená'] = row['cena_czk']; overwrite_czk += 1
            else:
                same_czk += 1

        if row['cena_eur'] is not None:
            if cur_eur is None:
                new_fields['Cena EUR doporučená'] = row['cena_eur']; new_eur += 1
            elif abs(float(cur_eur) - row['cena_eur']) > 0.005:
                new_fields['Cena EUR doporučená'] = row['cena_eur']; overwrite_eur += 1
            else:
                same_eur += 1

        if new_fields:
            patches.append({'id': rec['id'], 'fields': new_fields})
            if len(samples) < 5:
                samples.append((
                    row['kod'],
                    f'{cur_czk}→{new_fields.get("Cena CZK doporučená", "—")}',
                    f'{cur_eur}→{new_fields.get("Cena EUR doporučená", "—")}',
                ))

    print()
    print('=== Souhrn diffu ===')
    print(f'  CZK: nově={new_czk}  přepis={overwrite_czk}  beze změny={same_czk}')
    print(f'  EUR: nově={new_eur}  přepis={overwrite_eur}  beze změny={same_eur}')
    print(f'  Patches celkem (záznamy ke změně): {len(patches)}')

    if samples:
        print()
        print('=== Sample (max 5) ===')
        for kod, czk_diff, eur_diff in samples:
            print(f'  {kod}  CZK: {czk_diff}  |  EUR: {eur_diff}')

    if not patches:
        print('\nNic ke změně, končím.')
        return 0

    if args.dry_run:
        print('\nDRY-RUN — žádný zápis. Pro provedení spusť bez --dry-run.')
        return 0

    url = f'https://api.airtable.com/v0/{env["AIRTABLE_BASE_ID"]}/{urllib.parse.quote(env["AIRTABLE_TABLE_NAME"])}'
    success = 0
    for i in range(0, len(patches), 10):
        batch = patches[i:i+10]
        try:
            rj = patch_batch(url, env['AIRTABLE_TOKEN'], batch)
            success += len(rj.get('records', []))
            print(f'  batch {i//10+1}/{(len(patches)+9)//10}: OK ({len(rj.get("records", []))} rec)')
        except RuntimeError as e:
            print(f'  batch {i//10+1}: FAIL — {e}')
            break
        time.sleep(0.25)

    print(f'\nPushed: {success}/{len(patches)}')

    # Verifikace re-fetch
    print('\n=== Verifikace ===')
    at_records = fetch_at_records(env['AIRTABLE_TOKEN'], env['AIRTABLE_BASE_ID'], env['AIRTABLE_TABLE_NAME'])
    by_kod = {r['fields'].get('Kód Lakma'): r for r in at_records if r['fields'].get('Kód Lakma')}
    czk_match = eur_match = 0
    mismatch: list[tuple] = []
    for row in rows:
        rec = by_kod.get(row['kod'])
        if not rec:
            continue
        if row['cena_czk'] is not None:
            if rec['fields'].get('Cena CZK doporučená') is not None and abs(float(rec['fields']['Cena CZK doporučená']) - row['cena_czk']) <= 0.005:
                czk_match += 1
            else:
                mismatch.append((row['kod'], 'CZK', rec['fields'].get('Cena CZK doporučená'), row['cena_czk']))
        if row['cena_eur'] is not None:
            if rec['fields'].get('Cena EUR doporučená') is not None and abs(float(rec['fields']['Cena EUR doporučená']) - row['cena_eur']) <= 0.005:
                eur_match += 1
            else:
                mismatch.append((row['kod'], 'EUR', rec['fields'].get('Cena EUR doporučená'), row['cena_eur']))
    print(f'  CZK shoda: {czk_match}/{sum(1 for r in rows if r["cena_czk"] is not None and r["kod"] in by_kod)}')
    print(f'  EUR shoda: {eur_match}/{sum(1 for r in rows if r["cena_eur"] is not None and r["kod"] in by_kod)}')
    if mismatch:
        print(f'  Neshody ({len(mismatch)}):')
        for m in mismatch[:10]:
            print(f'    {m}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
