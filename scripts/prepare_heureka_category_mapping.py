"""
Připraví návrh Heureka CATEGORYTEXT pro produkty v Airtable Produkty_v2.

Skript nic nezapisuje do Airtable. Čte produkty, aplikuje mapovací config,
validuje cesty proti lokální referenci z oficiálního Heureka XML a vytvoří
review XLSX/CSV pro ruční kontrolu nebo následný import do polí:
- Heureka kategorie CZ
- Heureka kategorie SK

Nejdřív spusť:
  python scripts/update_heureka_category_reference.py

Výstup:
- data/heureka_categories/heureka_category_review_<date>.xlsx
- data/heureka_categories/heureka_category_import_<date>.csv
- data/heureka_categories/heureka_category_summary_<date>.json
"""
import argparse
import csv
import json
import re
import sys
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
ENV_PATH = ROOT / ".env"
CONFIG_PATH = ROOT / "config" / "heureka_category_mapping.json"
REFERENCE_DIR = ROOT / "data" / "reference" / "heureka"
OUTPUT_DIR = ROOT / "data" / "heureka_categories"

AT_FIELDS = [
    "Kód Lakma",
    "Název",
    "Web název CZ",
    "Web název SK",
    "Web produktová řada CZ",
    "Objem",
    "EAN KS",
    "Přidat do feedu",
    "Heureka kategorie CZ",
    "Heureka kategorie SK",
]


def normalize_path(value: str) -> str:
    parts = [" ".join(part.split()) for part in value.split("|")]
    return " | ".join(part for part in parts if part)


def normalize_text(value: str) -> str:
    return " ".join((value or "").lower().split())


def load_env() -> dict[str, str]:
    env = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.split("  ")[0].strip()
    return env


def fetch_records(token: str, base: str, table: str) -> list[dict]:
    records = []
    offset = None
    while True:
        params = [("pageSize", "100")]
        params.extend(("fields[]", field) for field in AT_FIELDS)
        if offset:
            params.append(("offset", offset))
        qs = urllib.parse.urlencode(params)
        url = f"https://api.airtable.com/v0/{base}/{urllib.parse.quote(table)}?{qs}"
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        records.extend(data.get("records", []))
        offset = data.get("offset")
        if not offset:
            break
    return records


def load_reference(market: str) -> set[str]:
    path = REFERENCE_DIR / f"heureka_{market}_categories.json"
    if not path.exists():
        raise FileNotFoundError(
            f"Chybí {path}. Nejdřív spusť scripts/update_heureka_category_reference.py"
        )
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {normalize_path(row["path"]) for row in payload["categories"]}


def category_paths(config: dict, category_key: str) -> tuple[str, str]:
    categories = config["category_keys"][category_key]
    return normalize_path(categories["cz"]), normalize_path(categories["sk"])


def match_rule(config: dict, fields: dict) -> dict | None:
    code = str(fields.get("Kód Lakma") or "")
    series = fields.get("Web produktová řada CZ") or ""
    name = normalize_text(
        " ".join(
            str(fields.get(field) or "")
            for field in ("Web název CZ", "Web název SK", "Název")
        )
    )

    exact = config.get("exact_code_rules", {}).get(code)
    if exact:
        return {**exact, "source": f"exact_code:{code}"}

    for rule in config.get("keyword_rules", []):
        if rule.get("series") and series not in rule["series"]:
            continue
        patterns = [normalize_text(value) for value in rule.get("contains_any", [])]
        if any(pattern and pattern in name for pattern in patterns):
            return {**rule, "source": f"keyword:{rule['id']}"}

    default = config.get("series_defaults", {}).get(series)
    if default:
        return {**default, "source": f"series:{series}"}
    return None


def build_rows(records: list[dict], config: dict, valid_cz: set[str], valid_sk: set[str]) -> list[dict]:
    excluded_codes = {str(code) for code in config.get("exclude_product_codes", [])}
    rows = []
    for record in records:
        fields = record.get("fields", {})
        code = str(fields.get("Kód Lakma") or "")
        if code in excluded_codes:
            rows.append(
                {
                    "record_id": record["id"],
                    "kod_lakma": code,
                    "web_nazev_cz": fields.get("Web název CZ") or fields.get("Název") or "",
                    "web_nazev_sk": fields.get("Web název SK") or "",
                    "rada": fields.get("Web produktová řada CZ") or "",
                    "objem": fields.get("Objem") or "",
                    "ean_ks": fields.get("EAN KS") or "",
                    "pridat_do_feedu": fields.get("Přidat do feedu") or "",
                    "aktualni_cz": normalize_path(fields.get("Heureka kategorie CZ") or ""),
                    "navrh_cz": "",
                    "validace_cz": "EXCLUDED",
                    "aktualni_sk": normalize_path(fields.get("Heureka kategorie SK") or ""),
                    "navrh_sk": "",
                    "validace_sk": "EXCLUDED",
                    "zmena_cz": "",
                    "zmena_sk": "",
                    "confidence": "",
                    "zdroj_pravidla": "excluded_product_code",
                    "poznamka": "Vyřazeno dle configu; Q Power privátka se nekategorizuje.",
                }
            )
            continue

        rule = match_rule(config, fields)
        if rule:
            cz_path, sk_path = category_paths(config, rule["category_key"])
            confidence = rule.get("confidence", "")
            source = rule.get("source", "")
            note = rule.get("note", "")
        else:
            cz_path = sk_path = ""
            confidence = ""
            source = "no_rule"
            note = "Chybí mapování pro řadu nebo název."

        current_cz = normalize_path(fields.get("Heureka kategorie CZ") or "")
        current_sk = normalize_path(fields.get("Heureka kategorie SK") or "")
        valid_cz_state = "OK" if cz_path in valid_cz else ("MISSING" if cz_path else "NO_MAPPING")
        valid_sk_state = "OK" if sk_path in valid_sk else ("MISSING" if sk_path else "NO_MAPPING")

        rows.append(
            {
                "record_id": record["id"],
                "kod_lakma": code,
                "web_nazev_cz": fields.get("Web název CZ") or fields.get("Název") or "",
                "web_nazev_sk": fields.get("Web název SK") or "",
                "rada": fields.get("Web produktová řada CZ") or "",
                "objem": fields.get("Objem") or "",
                "ean_ks": fields.get("EAN KS") or "",
                "pridat_do_feedu": fields.get("Přidat do feedu") or "",
                "aktualni_cz": current_cz,
                "navrh_cz": cz_path,
                "validace_cz": valid_cz_state,
                "aktualni_sk": current_sk,
                "navrh_sk": sk_path,
                "validace_sk": valid_sk_state,
                "zmena_cz": "ANO" if current_cz != cz_path else "NE",
                "zmena_sk": "ANO" if current_sk != sk_path else "NE",
                "confidence": confidence,
                "zdroj_pravidla": source,
                "poznamka": note,
            }
        )
    rows.sort(key=lambda row: (row["rada"], row["kod_lakma"]))
    return rows


def write_review_xlsx(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    columns = [
        ("Kód Lakma", "kod_lakma", 13),
        ("Web název CZ", "web_nazev_cz", 46),
        ("Web název SK", "web_nazev_sk", 46),
        ("Řada", "rada", 26),
        ("Objem", "objem", 10),
        ("EAN KS", "ean_ks", 16),
        ("Přidat do feedu", "pridat_do_feedu", 14),
        ("Aktuální CZ", "aktualni_cz", 58),
        ("Návrh CZ", "navrh_cz", 58),
        ("Validace CZ", "validace_cz", 12),
        ("Změna CZ", "zmena_cz", 10),
        ("Aktuální SK", "aktualni_sk", 58),
        ("Návrh SK", "navrh_sk", 58),
        ("Validace SK", "validace_sk", 12),
        ("Změna SK", "zmena_sk", 10),
        ("Confidence", "confidence", 11),
        ("Zdroj pravidla", "zdroj_pravidla", 28),
        ("Poznámka", "poznamka", 44),
    ]

    header_fill = PatternFill("solid", fgColor="0066B3")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    changed_fill = PatternFill("solid", fgColor="FFF4CC")
    ok_fill = PatternFill("solid", fgColor="E8F5E9")
    missing_fill = PatternFill("solid", fgColor="FFEBEE")
    excluded_fill = PatternFill("solid", fgColor="ECEFF1")
    thin = Side(style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (title, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(rows, start=2):
        state_fill = None
        if row["validace_cz"] == "EXCLUDED":
            state_fill = excluded_fill
        elif row["validace_cz"] == "OK" and row["validace_sk"] == "OK":
            state_fill = ok_fill
        elif row["validace_cz"] != "OK" or row["validace_sk"] != "OK":
            state_fill = missing_fill
        elif row["zmena_cz"] == "ANO" or row["zmena_sk"] == "ANO":
            state_fill = changed_fill

        for col_idx, (_, key, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.font = Font(name="Segoe UI", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if state_fill:
                cell.fill = state_fill

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ("Vygenerováno", datetime.now().isoformat(timespec="seconds")),
        ("Produktů celkem", len(rows)),
        ("Vyřazeno", sum(1 for row in rows if row["validace_cz"] == "EXCLUDED")),
        ("Validní CZ", sum(1 for row in rows if row["validace_cz"] == "OK")),
        ("Validní SK", sum(1 for row in rows if row["validace_sk"] == "OK")),
        ("Změna CZ", sum(1 for row in rows if row["zmena_cz"] == "ANO")),
        ("Změna SK", sum(1 for row in rows if row["zmena_sk"] == "ANO")),
    ]
    for row_idx, (key, value) in enumerate(summary_rows, start=1):
        summary.cell(row=row_idx, column=1, value=key).font = Font(name="Segoe UI", bold=True)
        summary.cell(row=row_idx, column=2, value=value)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 22

    wb.save(path)


def write_import_csv(path: Path, rows: list[dict]) -> None:
    fields = ["Kód Lakma", "Heureka kategorie CZ", "Heureka kategorie SK"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        writer.writeheader()
        for row in rows:
            if row["validace_cz"] != "OK" or row["validace_sk"] != "OK":
                continue
            writer.writerow(
                {
                    "Kód Lakma": row["kod_lakma"],
                    "Heureka kategorie CZ": row["navrh_cz"],
                    "Heureka kategorie SK": row["navrh_sk"],
                }
            )


def write_summary(path: Path, rows: list[dict]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "rows_total": len(rows),
        "excluded": sum(1 for row in rows if row["validace_cz"] == "EXCLUDED"),
        "valid_cz": sum(1 for row in rows if row["validace_cz"] == "OK"),
        "valid_sk": sum(1 for row in rows if row["validace_sk"] == "OK"),
        "changed_cz": sum(1 for row in rows if row["zmena_cz"] == "ANO"),
        "changed_sk": sum(1 for row in rows if row["zmena_sk"] == "ANO"),
        "validation_cz": dict(Counter(row["validace_cz"] for row in rows)),
        "validation_sk": dict(Counter(row["validace_sk"] for row in rows)),
        "rule_sources": dict(Counter(row["zdroj_pravidla"] for row in rows)),
        "low_confidence": [
            {
                "kod_lakma": row["kod_lakma"],
                "web_nazev_cz": row["web_nazev_cz"],
                "confidence": row["confidence"],
                "zdroj_pravidla": row["zdroj_pravidla"],
                "navrh_cz": row["navrh_cz"],
                "navrh_sk": row["navrh_sk"],
                "poznamka": row["poznamka"],
            }
            for row in rows
            if row["confidence"] != "" and float(row["confidence"]) < 0.8
        ],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def print_summary(rows: list[dict]) -> None:
    print(f"Produkty v review: {len(rows)}")
    print(f"  vyřazeno: {sum(1 for row in rows if row['validace_cz'] == 'EXCLUDED')}")
    print(f"  CZ validní: {sum(1 for row in rows if row['validace_cz'] == 'OK')}/{len(rows)}")
    print(f"  SK validní: {sum(1 for row in rows if row['validace_sk'] == 'OK')}/{len(rows)}")
    print(f"  změna CZ: {sum(1 for row in rows if row['zmena_cz'] == 'ANO')}")
    print(f"  změna SK: {sum(1 for row in rows if row['zmena_sk'] == 'ANO')}")

    invalid = [
        row
        for row in rows
        if row["validace_cz"] not in {"OK", "EXCLUDED"} or row["validace_sk"] not in {"OK", "EXCLUDED"}
    ]
    if invalid:
        print("\nNevalidní / bez mapování:")
        for row in invalid:
            print(f"  {row['kod_lakma']}: CZ={row['validace_cz']} SK={row['validace_sk']} {row['web_nazev_cz']}")


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=date.today().isoformat(), help="Suffix výstupních souborů.")
    args = parser.parse_args()

    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    valid_cz = load_reference("cz")
    valid_sk = load_reference("sk")
    env = load_env()

    records = fetch_records(env["AIRTABLE_TOKEN"], env["AIRTABLE_BASE_ID"], env["AIRTABLE_TABLE_NAME"])
    rows = build_rows(records, config, valid_cz, valid_sk)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    review_path = OUTPUT_DIR / f"heureka_category_review_{args.date}.xlsx"
    import_path = OUTPUT_DIR / f"heureka_category_import_{args.date}.csv"
    summary_path = OUTPUT_DIR / f"heureka_category_summary_{args.date}.json"

    write_review_xlsx(review_path, rows)
    write_import_csv(import_path, rows)
    write_summary(summary_path, rows)
    print_summary(rows)
    print(f"\nVýstupy:")
    print(f"  review: {review_path.relative_to(ROOT)}")
    print(f"  import: {import_path.relative_to(ROOT)}")
    print(f"  summary: {summary_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
