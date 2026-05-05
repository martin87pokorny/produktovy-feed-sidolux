"""
Generuje Excel pro výrobce s 143 produkty - výrobce doplní:
- Cenu bez DPH (CZK + EUR), sazba DPH se prefillne, cena s DPH se dopočítá formulí
- Země původu (prefilled "Polsko")
- MPN / Kód výrobce
- PAO (měsíce po otevření)
- Záruční dobu (měsíce)
- Volitelnou poznámku
"""
import json
import sys
import urllib.parse
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENV_PATH = Path(__file__).resolve().parent.parent / '.env'
OUTPUT_DIR = Path(__file__).resolve().parent.parent / 'data' / 'exchange'


def load_env():
    env = {}
    with ENV_PATH.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            k, v = line.split('=', 1)
            env[k.strip()] = v.split('  ')[0].strip()
    return env


def fetch_records(token, base, table):
    fields = ['Kód Lakma', 'Název', 'Web název CZ', 'EAN KS',
              'Web produktová řada CZ', 'Objem']
    records = []
    offset = None
    while True:
        params = [('pageSize', '100')]
        for f in fields:
            params.append(('fields[]', f))
        if offset:
            params.append(('offset', offset))
        qs = urllib.parse.urlencode(params)
        url = f'https://api.airtable.com/v0/{base}/{urllib.parse.quote(table)}?{qs}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        records.extend(data.get('records', []))
        offset = data.get('offset')
        if not offset:
            break
    return records


def main():
    sys.stdout.reconfigure(encoding='utf-8')
    env = load_env()
    records = fetch_records(env['AIRTABLE_TOKEN'], env['AIRTABLE_BASE_ID'], env['AIRTABLE_TABLE_NAME'])
    print(f'Stáhnuto {len(records)} produktů z Airtable.')

    records.sort(key=lambda r: (r['fields'].get('Web produktová řada CZ', ''), r['fields'].get('Kód Lakma', '')))

    wb = Workbook()
    # ============================ List 1: Ceník ============================
    ws = wb.active
    ws.title = 'Ceník'

    # (header_text, width, kind, prefilled_value)
    columns = [
        ('Kód Lakma',                    14, 'string',       None),
        ('Název produktu',               55, 'string',       None),
        ('EAN',                          16, 'string',       None),
        ('Řada',                         26, 'string',       None),
        ('Objem',                        10, 'string',       None),
        ('Cena bez DPH v CZK',           18, 'currency_czk', None),
        ('Sazba DPH CZ (%)',             14, 'percent',      21),
        ('Cena s DPH v CZK (auto)',      20, 'currency_czk', 'formula_czk'),
        ('Cena bez DPH v EUR',           18, 'currency_eur', None),
        ('Sazba DPH SK (%)',             14, 'percent',      23),
        ('Cena s DPH v EUR (auto)',      20, 'currency_eur', 'formula_eur'),
        ('Země původu',                  16, 'string',       'Polsko'),
        ('MPN / Kód výrobce',            18, 'string',       None),
        ('PAO (měs. po otevření)',       16, 'integer',      None),
        ('Záruční doba (měs.)',          16, 'integer',      None),
        ('Poznámka výrobce',             28, 'string',       None),
    ]

    HEADER_FILL = PatternFill('solid', fgColor='0066B3')
    HEADER_FONT = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN = Side(style='thin', color='B0BEC5')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Instrukce řádek 1-2
    last_col_letter = get_column_letter(len(columns))
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = ('Sidolux / Lakma — formulář pro doplnění cen a doplňujících údajů pro Heureka feed (CZ + SK). '
                'Vyplňte žluté sloupce. Modré (Cena s DPH) se dopočítají automaticky. Sloupce A–E jsou předvyplněné, neměňte je.')
    ws['A1'].font = Font(name='Segoe UI', size=10, italic=True, color='455A64')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 42

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = (f'Vygenerováno: {date.today().isoformat()} · Drogerie ZDE · '
                f'Počet produktů: {len(records)} · Párovací klíč: Kód Lakma (sloupec A)')
    ws['A2'].font = Font(name='Segoe UI', size=9, color='6B7989')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # Hlavička v řádku 4
    HDR_ROW = 4
    for col_idx, (name, width, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=HDR_ROW, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HDR_ROW].height = 42

    # Datové řádky
    DATA_ALIGN_LEFT = Alignment(vertical='center', wrap_text=True)
    DATA_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    LOCKED_FILL = PatternFill('solid', fgColor='F5F7FA')   # předvyplněné
    INPUT_FILL = PatternFill('solid', fgColor='FFF8E8')    # výrobce vyplní
    AUTO_FILL = PatternFill('solid', fgColor='E3F2FD')     # automaticky dopočítané
    PREFILL_FILL = PatternFill('solid', fgColor='E8F5E9')  # prefilled (výrobce může upravit)

    # Mapování indexu sloupce -> typ buňky pro fill
    # 1-5 (A-E): locked (předvyplněné neměnné z naší DB)
    # 6 (F):    input - cena bez DPH CZK
    # 7 (G):    prefill - sazba DPH CZ (21)
    # 8 (H):    auto - formula CZK
    # 9 (I):    input - cena bez DPH EUR
    # 10 (J):   prefill - sazba DPH SK (23)
    # 11 (K):   auto - formula EUR
    # 12 (L):   prefill - země původu Polsko
    # 13 (M):   input - MPN
    # 14 (N):   input - PAO
    # 15 (O):   input - záruka
    # 16 (P):   input - poznámka

    LOCKED_COLS = {1, 2, 3, 4, 5}
    AUTO_COLS = {8, 11}
    PREFILL_COLS = {7, 10, 12}

    for i, r in enumerate(records, start=HDR_ROW + 1):
        f = r['fields']
        kod = str(f.get('Kód Lakma') or '')
        nazev = f.get('Web název CZ') or f.get('Název') or ''
        ean = str(f.get('EAN KS') or '')
        rada = f.get('Web produktová řada CZ') or ''
        objem = f.get('Objem') or ''

        for col_idx, (_, _, kind, prefilled) in enumerate(columns, start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.border = BORDER
            cell.font = Font(name='Segoe UI', size=10)
            cell.alignment = DATA_ALIGN_CENTER if col_idx in (1, 3, 5, 14, 15) else DATA_ALIGN_LEFT

            # fill barvy
            if col_idx in LOCKED_COLS:
                cell.fill = LOCKED_FILL
            elif col_idx in AUTO_COLS:
                cell.fill = AUTO_FILL
            elif col_idx in PREFILL_COLS:
                cell.fill = PREFILL_FILL
            else:
                cell.fill = INPUT_FILL

            # hodnota
            if col_idx == 1:
                cell.value = kod
            elif col_idx == 2:
                cell.value = nazev
            elif col_idx == 3:
                cell.value = ean
            elif col_idx == 4:
                cell.value = rada
            elif col_idx == 5:
                cell.value = objem
            elif prefilled == 'formula_czk':
                cell.value = f'=F{i}*(1+G{i}/100)'
            elif prefilled == 'formula_eur':
                cell.value = f'=I{i}*(1+J{i}/100)'
            elif prefilled is not None:
                cell.value = prefilled

            # number format
            if kind == 'currency_czk':
                cell.number_format = '#,##0.00 "Kč"'
            elif kind == 'currency_eur':
                cell.number_format = '#,##0.00 "€"'
            elif kind == 'percent':
                cell.number_format = '0"%"'
            elif kind == 'integer':
                cell.number_format = '0'

    # Zmrazit hlavičku + první sloupec
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=2)

    # ============================ List 2: Pokyny ============================
    ws2 = wb.create_sheet('Pokyny')
    ws2.column_dimensions['A'].width = 110
    pokyny = [
        ('Pokyny pro vyplnění formuláře', Font(name='Segoe UI', size=14, bold=True, color='0066B3')),
        ('', None),
        ('Barevné označení sloupců:', Font(name='Segoe UI', size=11, bold=True)),
        ('   • šedivé (A–E): naše data, neměňte je. Slouží jako párovací klíč a reference.',
         Font(name='Segoe UI', size=10, color='455A64')),
        ('   • žluté: tady doplňte hodnoty.',
         Font(name='Segoe UI', size=10, color='455A64')),
        ('   • zelené: prefilled (předvyplněno) — můžete upravit, pokud má daný produkt jinou hodnotu.',
         Font(name='Segoe UI', size=10, color='455A64')),
        ('   • modré: automaticky dopočítané (cena s DPH = cena bez DPH × (1 + sazba/100)). Neměňte.',
         Font(name='Segoe UI', size=10, color='455A64')),
        ('', None),
        ('1. Cena bez DPH v CZK (sloupec F) a EUR (sloupec I) — povinné pro každý produkt.',
         Font(name='Segoe UI', size=11)),
        ('   Doporučená maloobchodní cena pro Heureka feed (CZ + SK). Sidolux není e-shop, je to pouze referenční údaj.',
         Font(name='Segoe UI', size=10, italic=True, color='455A64')),
        ('', None),
        ('2. Sazba DPH (sloupce G, J) — prefilled na 21 % (CZ) a 23 % (SK). Pokud má některý produkt jinou sazbu, přepište.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('3. Cena s DPH (sloupce H, K) — automaticky dopočítané, NEMĚŇTE je. Excel počítá formulí.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('4. Země původu (sloupec L) — prefilled "Polsko". Pokud je produkt vyráběn jinde (např. Mr. Teppich v jiné zemi), přepište.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('5. MPN / Kód výrobce (sloupec M) — pokud máte interní kód odlišný od Kódu Lakma, doplňte.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('6. PAO (sloupec N) — Period After Opening v měsících. Symbol "12M" → vyplníte 12. Pokud produkt PAO nemá, nechte prázdné.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('7. Záruční doba (sloupec O) — v měsících. Pokud máte různou pro CZ vs SK, uveďte CZ a poznamenejte do sloupce P.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('8. Pokud k některému produktu hodnoty nejsou (vzorek, výprodej, ukončený), nechte buňky prázdné a doplňte poznámku do P.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('9. Vyplněný soubor pošlete zpět ve formátu XLSX. Importujeme automaticky podle Kódu Lakma.',
         Font(name='Segoe UI', size=11)),
        ('', None),
        ('Kontakt: honza@drogeriezde.cz',
         Font(name='Segoe UI', size=10, color='6B7989')),
    ]
    for row_idx, (text, font) in enumerate(pokyny, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    today = date.today().isoformat()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUTPUT_DIR / f'cenik_pro_vyrobce_{today}.xlsx'
    wb.save(out)
    print(f'OK: {out} ({out.stat().st_size / 1024:.1f} kB)')
    print(f'\nSloupce: {len(columns)} (A-{last_col_letter})')


if __name__ == '__main__':
    main()
